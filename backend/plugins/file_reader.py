import os
import io
import logging
from typing import Dict, Any, Union

# Importa as bibliotecas necessárias para diferentes tipos de ficheiro
try:
    import pypdf # Para PDFs
except ImportError:
    logging.warning("pypdf não está instalado. Não será possível ler ficheiros PDF.")
    pypdf = None

try:
    import docx # Para Word (.docx)
except ImportError:
    logging.warning("python-docx não está instalado. Não será possível ler ficheiros DOCX.")
    docx = None

try:
    import openpyxl # Para Excel (.xlsx)
except ImportError:
    logging.warning("openpyxl não está instalado. Não será possível ler ficheiros XLSX.")
    openpyxl = None

# Configura o logging para este módulo
logger = logging.getLogger(__name__)

class FileReader:
    """
    Uma classe para ler e extrair conteúdo de texto de vários tipos de ficheiro,
    incluindo TXT, PDF, DOCX e XLSX.
    """
    def __init__(self):
        """
        Inicializa a classe FileReader.
        """
        pass

    def _read_text(self, file_path: str) -> Dict[str, str]:
        """Lê o conteúdo de um ficheiro de texto (.txt)."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return {"status": "success", "content": content}
        except FileNotFoundError:
            return {"status": "error", "message": f"Ficheiro não encontrado: {file_path}"}
        except Exception as e:
            return {"status": "error", "message": f"Erro ao ler ficheiro TXT: {str(e)}"}

    def _read_pdf(self, file_path: str) -> Dict[str, str]:
        """Extrai texto de um ficheiro PDF (.pdf)."""
        if not pypdf:
            return {"status": "error", "message": "A biblioteca 'pypdf' não está instalada. Não é possível ler ficheiros PDF."}
        try:
            text_content = []
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                for page_num in range(len(reader.pages)):
                    text_content.append(reader.pages[page_num].extract_text())
            return {"status": "success", "content": "\n".join(text_content)}
        except FileNotFoundError:
            return {"status": "error", "message": f"Ficheiro não encontrado: {file_path}"}
        except pypdf.errors.PdfReadError as e:
            return {"status": "error", "message": f"Erro ao ler ficheiro PDF (pode estar corrompido ou protegido): {str(e)}"}
        except Exception as e:
            return {"status": "error", "message": f"Erro inesperado ao ler ficheiro PDF: {str(e)}"}

    def _read_docx(self, file_path: str) -> Dict[str, str]:
        """Extrai texto de um ficheiro Word (.docx)."""
        if not docx:
            return {"status": "error", "message": "A biblioteca 'python-docx' não está instalada. Não é possível ler ficheiros DOCX."}
        try:
            document = docx.Document(file_path)
            text_content = [paragraph.text for paragraph in document.paragraphs]
            return {"status": "success", "content": "\n".join(text_content)}
        except FileNotFoundError:
            return {"status": "error", "message": f"Ficheiro não encontrado: {file_path}"}
        except Exception as e:
            return {"status": "error", "message": f"Erro ao ler ficheiro DOCX: {str(e)}"}

    def _read_xlsx(self, file_path: str) -> Dict[str, str]:
        """Extrai texto de um ficheiro Excel (.xlsx)."""
        if not openpyxl:
            return {"status": "error", "message": "A biblioteca 'openpyxl' não está instalada. Não é possível ler ficheiros XLSX."}
        try:
            workbook = openpyxl.load_workbook(file_path)
            text_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"--- Folha: {sheet_name} ---")
                for row in sheet.iter_rows():
                    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                    text_content.append("\t".join(row_values)) # Usa tab para separar colunas
                text_content.append("\n")
            return {"status": "success", "content": "\n".join(text_content)}
        except FileNotFoundError:
            return {"status": "error", "message": f"Ficheiro não encontrado: {file_path}"}
        except Exception as e:
            return {"status": "error", "message": f"Erro ao ler ficheiro XLSX: {str(e)}"}

    def read_file(self, params: Dict[str, str]) -> Dict[str, str]:
        """
        Lê o conteúdo de um ficheiro especificado pelo caminho.
        Determina o tipo de ficheiro pela extensão e chama o leitor apropriado.

        Args:
            params (Dict[str, str]): Um dicionário contendo:
                - 'file_path' (str): O caminho completo para o ficheiro a ser lido.

        Returns:
            Dict[str, str]: Um dicionário com 'status' ('success' ou 'error') e 'content'/'message'.
        """
        file_path = params.get('file_path')

        if not file_path:
            return {"status": "error", "message": "O caminho do ficheiro é obrigatório."}
        
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"O ficheiro não existe no caminho especificado: {file_path}"}

        file_extension = os.path.splitext(file_path)[1].lower()

        readers = {
            '.txt': self._read_text,
            '.pdf': self._read_pdf,
            '.docx': self._read_docx,
            '.xlsx': self._read_xlsx,
        }

        reader_func = readers.get(file_extension)

        if reader_func:
            logger.info(f"A ler ficheiro '{file_path}' com o leitor para '{file_extension}'.")
            return reader_func(file_path)
        else:
            return {"status": "error", "message": f"Tipo de ficheiro '{file_extension}' não suportado."}

def register() -> Dict[str, Any]:
    """
    Registra a classe FileReader como uma ferramenta para o TaskEngine.
    Define a ação 'read_file'.

    Returns:
        Dict[str, Any]: Um dicionário que descreve a ferramenta e suas ações.
    """
    file_reader_instance = FileReader()
    return {
        "name": "file_reader", # Este é o nome que aparecerá na lista de plugins carregados
        "description": "Ferramenta para ler e extrair conteúdo de texto de ficheiros (TXT, PDF, DOCX, XLSX).",
        "actions": {
            "read_file": {
                "description": "Lê o conteúdo de texto de um ficheiro especificado pelo caminho.",
                "parameters": {
                    "file_path": "string (caminho completo para o ficheiro)"
                },
                "execute": file_reader_instance.read_file
            }
        }
    }

# Exemplo de uso interativo (para testes diretos do plugin)
if __name__ == "__main__":
    tool_info = register()
    file_reader_instance = tool_info["actions"]["read_file"]["execute"].__self__

    print("--- Teste de Leitura de Ficheiros ---")

    # Crie alguns ficheiros de teste para experimentar
    test_dir = "test_files"
    os.makedirs(test_dir, exist_ok=True)
    with open(os.path.join(test_dir, "exemplo.txt"), "w", encoding="utf-8") as f:
        f.write("Este é um ficheiro de texto de exemplo.\nContém algumas linhas de texto.")

    # Exemplo de leitura de TXT
    print("\nLendo exemplo.txt:")
    result_txt = file_reader_instance.read_file({"file_path": os.path.join(test_dir, "exemplo.txt")})
    print(result_txt)

    # Exemplo de ficheiro não existente
    print("\nLendo ficheiro_inexistente.txt:")
    result_non_existent = file_reader_instance.read_file({"file_path": "ficheiro_inexistente.txt"})
    print(result_non_existent)

